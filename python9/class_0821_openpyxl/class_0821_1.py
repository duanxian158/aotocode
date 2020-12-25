#引入模块import from...import
#一层一层拨开 放在lib下面 我不管
#命名规范：数字 字母 下划线组成  不能以数字开头
#类名 首字母大写驼峰命名
#函数 模块名 变量名 小字书字母不同字母用下划线隔开 方便阅读
#解决我们测试数据的问题
#测试用例：数据库 Excel
#怎么操作Excel
#打开工作簿---找到表单---定位到你要的单元格
#openpyxl：同时读写 专门来操作Excel
## xlrd:只能读 xlwt:只能写
from openpyxl import load_workbook
#1：打开工作簿 必须是已经存在的工作簿
#openpyxl 操作Excel 必须是xlsx结尾 不能是xls结尾 否则报错
wb=load_workbook('python9.xlsx')

#2:定位表单 sheet
sheet=wb['Sheet1']
#3:定位单元格 获取单元格里面的数据
#根据行列去定位 坐标 去定位元素
print(sheet.cell(2,1).value)
#如果你获取的值 返回的是None 没有数据 要么就是没取到
sheet.cell(2,1).value='登录成功'
#如果有数据更新，那么就一定要记得保存文件
wb.save('python9.xlsx')
#坑二：如果涉及到写入数据，
# 那么Excel一定不能是打开状态？

