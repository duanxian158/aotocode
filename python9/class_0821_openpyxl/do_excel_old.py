
from openpyxl import load_workbook
wb=load_workbook('python9.xlsx')
sheet=wb['Sheet1']
#读取所有的数据
#1:先完成第一行数据的读取
#2：再来想办法 完成第二行数据的读取
# all_data=[]
# max_col=sheet.max_column#获取数据的最大列
# max_r=sheet.max_row#获取最大行
#
# for i in range(1,max_r+1):#控制行
#     row_data = []
#     for j in range(1, max_col+1):#控制列
#         row_data.append(sheet.cell(i, j).value)
#     all_data.append(row_data)
# print(all_data)
 #数据从Excelj里面读取出来是什么类型
data=sheet.cell(3,5).value
print(data)
print(type(data))