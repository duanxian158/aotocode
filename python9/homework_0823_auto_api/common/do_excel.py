from openpyxl import load_workbook
#请把每一行数据存到字典里面  其中key为对应的表头
#所有字典存到一个大列表里面
class DoExcel:
    def __init__(self,file_name,sheet_name):#初始化函数 Excel文件名 表单名
        self.file_name=file_name
        self.sheet_name=sheet_name
    def get_title(self):#获取第一行的表头字段
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        title=[]#存储表头
        for i in range(1,sheet.max_column+1):
            title.append(sheet.cell(1,i).value)
        return title
    def do_excel(self,mode,case_id_list):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        #获取title
        title=self.get_title()
        all_data=[]#大列表
        max_col=sheet.max_column#获取数据的最大列
        max_r=sheet.max_row#获取最大行

        for i in range(2,max_r+1):#控制行
            row_data = {}#每一行数据存到一个字典里
            for j in range(1, max_col+1):#控制列
                # key=title[j-1]#j-1 是第一列 对应的title
                # value=sheet.cell(i,j).value
                # print('key:{0}'.format(key))
                # print('value:{0}'.format(value))
                row_data[title[j-1]]=sheet.cell(i,j).value
            all_data.append(row_data)
        if mode=='1':#如果等于1 就跑所有的用例
            final_data=all_data
        else:#否则就执行case_id_list里面的用例
            final_data = []  # 最终的测试用例数据 存到这个列表里
            for item in all_data:#对所有的数据进行遍历
                if item['case_id'] in eval(case_id_list):#去判断case_id是否相等
                    final_data.append(item)#如果相等的话，那么就把测试数据加入到final_data列表里

        return final_data#返回最终的数据
if __name__ == '__main__':
    from homework_0823_auto_api.conf import read_path
    result=DoExcel(read_path.test_data_path,'Sheet2').do_excel('1','[1]')
    print(result)
