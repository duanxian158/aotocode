from openpyxl import load_workbook
from common import pro_path
from common.read_config import  ReadConfig
class DoExcel:
    def __init__(self, file_path, sheet_name):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def do_excel(self):

        flag = ReadConfig().read_config(pro_path.conf_path, 'TESTCASE', 'flag')
        case_id_list = ReadConfig().read_config(pro_path.conf_path, 'TESTCASE', 'case_id_list')
        wb = load_workbook(self.file_path)#打开工作簿
        sheet = wb[self.sheet_name]#获取表单
        tel = self.get_tel()#从Excel里获取初始化手机号
        self.update_tel(tel+1)#获取完毕之后  更新初始化的手机号码

        self.used_tel(tel)#调用函数  存储手机号码

        test_data=[] #所有的测试数据放到一个列表里
        for i in range(2,sheet.max_row+1):
            sub_data = {}#每一行数据存入到字典里
            sub_data['CaseId'] = sheet.cell(i,1).value
            sub_data['Title'] = sheet.cell(i, 2).value
            sub_data['Method'] = sheet.cell(i, 3).value
            sub_data['URL'] = sheet.cell(i, 4).value
            sub_data['Param'] = sheet.cell(i, 5).value #请求参数
            if sub_data['Param'].find('${tel}')!=-1:#find函数  如果存在，
                sub_data['Param'].replace('${tel}',str(tel))

            sub_data['ExpectedResult'] = sheet.cell(i, 6).value
            test_data.append(sub_data)#添加数据到列表里面
        #根据配置文件的开关来决定运行哪些测试用例
        if flag=='on':#意味着执行所有的用例
            final_data = test_data
        else:#执行指定列表里面的数据
            final_data=[]#用来存储最后要执行的用例
            for i in eval(case_id_list):
                final_data.append(test_data[i-1])#注意这里的写法？
        return final_data#返回最终的数据

    def write_back(self,row,col,new_value):#写回数据
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]

        sheet.cell(row,col).value = new_value
        wb.save(self.file_path)

    def used_tel(self,used_tel):#存储已经使用过的电话号码
        wb = load_workbook(self.file_path)
        sheet = wb['used']
        sheet.cell(sheet.max_row+1,1).value = used_tel
        wb.save(self.file_path)

    def get_tel(self):#获取写在Excel里面的初始化手机号
        wb = load_workbook(self.file_path)
        sheet = wb['tel']
        tel = sheet.cell(1,1).value
        return tel
    def update_tel(self,new_tel):#更新初始化手机号码函数
        wb = load_workbook(self.file_path)
        sheet = wb['tel']
        sheet.cell(1, 1).value = new_tel
        wb.save(self.file_path)

if __name__ == '__main__':
    test_data = DoExcel('D:/python3.6/code/python9/AUTO_API_TEST_1/test_data/test_1.xlsx','Sheet2').do_excel()
    print('获取到的测试数据是：{0}'.format(test_data))
    tel = DoExcel('D:/python3.6/code/python9/AUTO_API_TEST_1/test_data/test_1.xlsx','Sheet2').get_tel(2)
    print('获取到的tel:{0}'.format(tel))

